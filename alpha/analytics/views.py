"""
Analytics Views - COMPLETE VERSION WITH DAILY AND MONTHLY SUPPORT
Place this at: alpha/analytics/views.py
"""
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, View
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Sum, Q, F, Avg
from django.db.models.functions import TruncMonth, TruncDay
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import json

from alpha.visits.models import Visit
from alpha.appointments.models import Appointment
from alpha.clients.models import Client

# For Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ═══════════════════════════════════════════════════════════════
# MAIN DASHBOARD VIEW - SUPPORTS DAILY AND MONTHLY
# ═══════════════════════════════════════════════════════════════

class AnalyticsDashboardView(LoginRequiredMixin, TemplateView):
    """Main analytics dashboard with charts and statistics - SUPPORTS DAILY AND MONTHLY"""
    template_name = 'analytics/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get view mode (daily or monthly) from URL
        mode = self.request.GET.get('mode', 'monthly')  # Default to monthly
        context['mode'] = mode
        
        # Get today's date
        today = timezone.now().date()
        context['today_str'] = today.strftime('%Y-%m-%d')
        
        if mode == 'daily':
            # ═══════════════════════════════════════════════════════
            # DAILY MODE
            # ═══════════════════════════════════════════════════════
            selected_date_str = self.request.GET.get('date')
            if selected_date_str:
                try:
                    selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
                except ValueError:
                    selected_date = today
            else:
                selected_date = today
            
            context['selected_date'] = selected_date
            context['selected_date_str'] = selected_date.strftime('%Y-%m-%d')
            context['date_display'] = selected_date.strftime('%A, %B %d, %Y')
            context['is_today'] = selected_date == today
            
            # Date range for the selected day
            start_date = datetime.combine(selected_date, datetime.min.time())
            end_date = datetime.combine(selected_date, datetime.max.time())
            start_date = timezone.make_aware(start_date)
            end_date = timezone.make_aware(end_date)
            
            # Get daily statistics
            context['stats'] = self._get_daily_stats(start_date, end_date)
            
            # Pass available_months for template compatibility
            context['available_months'] = self._get_available_months()
            context['selected_month'] = today.strftime('%Y-%m')
            
        else:
            # ═══════════════════════════════════════════════════════
            # MONTHLY MODE (Your original logic)
            # ═══════════════════════════════════════════════════════
            selected_month = self.request.GET.get('month')
            if selected_month:
                try:
                    selected_date = datetime.strptime(selected_month, '%Y-%m')
                except ValueError:
                    selected_date = timezone.now()
            else:
                selected_date = timezone.now()
            
            # Calculate date range for the selected month
            start_date = selected_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if start_date.month == 12:
                end_date = start_date.replace(year=start_date.year + 1, month=1)
            else:
                end_date = start_date.replace(month=start_date.month + 1)
            
            context['selected_month'] = selected_date.strftime('%Y-%m')
            context['month_name'] = selected_date.strftime('%B %Y')
            context['date_display'] = selected_date.strftime('%B %Y')
            
            # Get available months (last 12 months)
            context['available_months'] = self._get_available_months()
            
            # Get monthly statistics
            context['stats'] = self._get_monthly_stats(start_date, end_date)
            
            # Pass selected_date_str for template compatibility
            context['selected_date_str'] = today.strftime('%Y-%m-%d')
        
        return context
    
    def _get_available_months(self):
        """Get list of last 12 months for dropdown"""
        months = []
        current = timezone.now().replace(day=1)
        for i in range(12):
            months.append({
                'value': current.strftime('%Y-%m'),
                'label': current.strftime('%B %Y')
            })
            if current.month == 1:
                current = current.replace(year=current.year - 1, month=12)
            else:
                current = current.replace(month=current.month - 1)
        return months
    
    def _get_daily_stats(self, start_date, end_date):
        """Calculate key statistics for a single day"""
        
        # Filter visits and appointments for the day
        visits = Visit.objects.filter(
            appointment__start__gte=start_date,
            appointment__start__lt=end_date
        )
        
        appointments = Appointment.objects.filter(
            start__gte=start_date,
            start__lt=end_date
        )
        
        # Calculate statistics (SAME FORMAT as monthly)
        stats = {
            'total_visits': visits.count(),
            'total_appointments': appointments.count(),
            'new_clients': appointments.values('client').distinct().count(),  # Unique clients that day
            
            # Revenue calculations
            'revenue': visits.aggregate(total=Sum('charge_amount'))['total'] or Decimal('0.00'),
            'paid': visits.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0.00'),
            'pending': (visits.aggregate(total=Sum('charge_amount'))['total'] or Decimal('0.00')) - 
                      (visits.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0.00')),
            
            # Payment status counts
            'fully_paid': visits.filter(
                paid_amount__gte=F('charge_amount'),
                charge_amount__gt=0
            ).count(),
            'partially_paid': visits.filter(
                paid_amount__gt=0,
                paid_amount__lt=F('charge_amount')
            ).count(),
            'unpaid': visits.filter(
                Q(paid_amount=0) | Q(paid_amount__isnull=True)
            ).count(),
            
            # Appointment status counts
            'completed': appointments.filter(status='completed').count(),
            'booked': appointments.filter(status='booked').count(),
            'cancelled': appointments.filter(status='cancelled').count(),
            'no_show': appointments.filter(status='no_show').count(),
        }
        
        return stats
    
    def _get_monthly_stats(self, start_date, end_date):
        """Calculate key statistics for the month"""
        
        # Filter visits and appointments for the month
        visits = Visit.objects.filter(
            appointment__start__gte=start_date,
            appointment__start__lt=end_date
        )
        
        appointments = Appointment.objects.filter(
            start__gte=start_date,
            start__lt=end_date
        )
        
        # Calculate statistics
        stats = {
            'total_visits': visits.count(),
            'total_appointments': appointments.count(),
            'new_clients': appointments.values('client').distinct().count(),  # Unique clients that month
            
            # Revenue calculations
            'revenue': visits.aggregate(total=Sum('charge_amount'))['total'] or Decimal('0.00'),
            'paid': visits.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0.00'),
            'pending': (visits.aggregate(total=Sum('charge_amount'))['total'] or Decimal('0.00')) - 
                      (visits.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0.00')),
            
            # Payment status counts
            'fully_paid': visits.filter(
                paid_amount__gte=F('charge_amount'),
                charge_amount__gt=0
            ).count(),
            'partially_paid': visits.filter(
                paid_amount__gt=0,
                paid_amount__lt=F('charge_amount')
            ).count(),
            'unpaid': visits.filter(
                Q(paid_amount=0) | Q(paid_amount__isnull=True)
            ).count(),
            
            # Appointment status counts
            'completed': appointments.filter(status='completed').count(),
            'booked': appointments.filter(status='booked').count(),
            'cancelled': appointments.filter(status='cancelled').count(),
            'no_show': appointments.filter(status='no_show').count(),
        }
        
        return stats


# ═══════════════════════════════════════════════════════════════
# CHART DATA API - SUPPORTS DAILY AND MONTHLY
# ═══════════════════════════════════════════════════════════════

class AnalyticsDataAPIView(LoginRequiredMixin, View):
    """API view for chart data - UPDATED TO SUPPORT DAILY AND MONTHLY"""
    
    def get(self, request):
        chart_type = request.GET.get('type')
        month = request.GET.get('month')
        date_str = request.GET.get('date')  # For daily mode
        
        # Determine if this is daily or monthly request
        if date_str:
            # DAILY MODE - use single date
            try:
                selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                selected_date = timezone.now().date()
            
            start_datetime = timezone.make_aware(datetime.combine(selected_date, datetime.min.time()))
            end_datetime = timezone.make_aware(datetime.combine(selected_date, datetime.max.time()))
            
            # Route to daily chart methods
            if chart_type == 'hourly_revenue':
                return self._get_hourly_revenue_data(start_datetime, end_datetime)
            elif chart_type == 'visits':
                return self._get_visits_data(start_datetime, end_datetime)
            elif chart_type == 'services':
                return self._get_services_data(start_datetime, end_datetime)
            elif chart_type == 'staff':
                return self._get_staff_data(start_datetime, end_datetime)
            elif chart_type == 'rooms':
                return self._get_rooms_data(start_datetime, end_datetime)
            elif chart_type == 'machines':
                return self._get_machines_data(start_datetime, end_datetime)
        
        else:
            # MONTHLY MODE - use month
            if not month:
                month = timezone.now().strftime('%Y-%m')
            
            try:
                selected_date = datetime.strptime(month, '%Y-%m')
            except ValueError:
                selected_date = timezone.now()
            
            start_date = selected_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if start_date.month == 12:
                end_date = start_date.replace(year=start_date.year + 1, month=1)
            else:
                end_date = start_date.replace(month=start_date.month + 1)
            
            # Route to monthly chart methods
            if chart_type == 'daily_revenue':
                return self._get_daily_revenue_data(start_date, end_date)
            elif chart_type == 'visits':
                return self._get_visits_data(start_date, end_date)
            elif chart_type == 'services':
                return self._get_services_data(start_date, end_date)
            elif chart_type == 'staff':
                return self._get_staff_data(start_date, end_date)
            elif chart_type == 'rooms':
                return self._get_rooms_data(start_date, end_date)
            elif chart_type == 'machines':
                return self._get_machines_data(start_date, end_date)
        
        return JsonResponse({'error': 'Invalid chart type'}, status=400)
    
    # ═══════════════════════════════════════════════════════════════
    # DAILY CHART - Hourly Revenue
    # ═══════════════════════════════════════════════════════════════
    
    def _get_hourly_revenue_data(self, start_datetime, end_datetime):
        """Get hourly revenue breakdown for a single day"""
        labels = []
        revenue_data = []
        appointments_data = []
        
        # Loop through hours 7 AM to 10 PM
        for hour in range(7, 23):
            hour_start = start_datetime.replace(hour=hour, minute=0, second=0)
            hour_end = start_datetime.replace(hour=hour, minute=59, second=59)
            
            # Get revenue for this hour
            visits = Visit.objects.filter(
                appointment__start__gte=hour_start,
                appointment__start__lte=hour_end
            )
            
            hour_revenue = visits.aggregate(
                total=Sum('paid_amount')
            )['total'] or Decimal('0.00')
            
            # Get appointments for this hour
            hour_appointments = Appointment.objects.filter(
                start__gte=hour_start,
                start__lte=hour_end
            ).count()
            
            labels.append(f'{hour:02d}:00')
            revenue_data.append(float(hour_revenue))
            appointments_data.append(hour_appointments)
        
        return JsonResponse({
            'labels': labels,
            'datasets': [
                {
                    'label': 'Revenue (€)',
                    'data': revenue_data,
                    'borderColor': 'rgb(75, 192, 192)',
                    'backgroundColor': 'rgba(75, 192, 192, 0.2)',
                    'tension': 0.4
                },
                {
                    'label': 'Appointments',
                    'data': appointments_data,
                    'borderColor': 'rgb(54, 162, 235)',
                    'backgroundColor': 'rgba(54, 162, 235, 0.2)',
                    'tension': 0.4
                }
            ]
        })
    
    # ═══════════════════════════════════════════════════════════════
    # MONTHLY CHART - Daily Revenue
    # ═══════════════════════════════════════════════════════════════
    
    def _get_daily_revenue_data(self, start_date, end_date):
        """Daily revenue trend for the month"""
        visits_by_day = Visit.objects.filter(
            appointment__start__gte=start_date,
            appointment__start__lt=end_date
        ).annotate(
            day=TruncDay('appointment__start')
        ).values('day').annotate(
            total_revenue=Sum('paid_amount'),
            total_charged=Sum('charge_amount')
        ).order_by('day')
        
        labels = []
        revenue_data = []
        charged_data = []
        
        current_date = start_date.date()
        end_date_only = end_date.date()
        
        data_dict = {item['day'].date(): item for item in visits_by_day}
        
        while current_date < end_date_only:
            labels.append(current_date.strftime('%-d'))
            
            if current_date in data_dict:
                revenue_data.append(float(data_dict[current_date]['total_revenue'] or 0))
                charged_data.append(float(data_dict[current_date]['total_charged'] or 0))
            else:
                revenue_data.append(0)
                charged_data.append(0)
            
            current_date += timedelta(days=1)
        
        return JsonResponse({
            'labels': labels,
            'datasets': [
                {
                    'label': 'Paid Revenue (€)',
                    'data': revenue_data,
                    'borderColor': 'rgb(75, 192, 192)',
                    'backgroundColor': 'rgba(75, 192, 192, 0.2)',
                    'tension': 0.4
                },
                {
                    'label': 'Charged (€)',
                    'data': charged_data,
                    'borderColor': 'rgb(255, 206, 86)',
                    'backgroundColor': 'rgba(255, 206, 86, 0.2)',
                    'tension': 0.4,
                    'borderDash': [5, 5]
                }
            ]
        })
    
    # ═══════════════════════════════════════════════════════════════
    # SHARED CHARTS - Work for both daily and monthly
    # ═══════════════════════════════════════════════════════════════
    
    def _get_visits_data(self, start_date, end_date):
        """Payment status breakdown"""
        visits = Visit.objects.filter(
            appointment__start__gte=start_date,
            appointment__start__lt=end_date
        )
        
        fully_paid = visits.filter(
            paid_amount__gte=F('charge_amount'),
            charge_amount__gt=0
        ).count()
        
        partially_paid = visits.filter(
            paid_amount__gt=0,
            paid_amount__lt=F('charge_amount')
        ).count()
        
        unpaid = visits.filter(
            Q(paid_amount=0) | Q(paid_amount__isnull=True)
        ).count()
        
        return JsonResponse({
            'labels': ['Fully Paid', 'Partially Paid', 'Unpaid'],
            'datasets': [{
                'data': [fully_paid, partially_paid, unpaid],
                'backgroundColor': [
                    'rgba(75, 192, 192, 0.8)',
                    'rgba(255, 206, 86, 0.8)',
                    'rgba(255, 99, 132, 0.8)'
                ],
                'borderWidth': 2
            }]
        })
    
    def _get_services_data(self, start_date, end_date):
        """Top services by revenue"""
        services = Visit.objects.filter(
            appointment__start__gte=start_date,
            appointment__start__lt=end_date
        ).values(
            'appointment__service__name'
        ).annotate(
            total_revenue=Sum('paid_amount')
        ).order_by('-total_revenue')[:10]
        
        labels = [s['appointment__service__name'] for s in services]
        data = [float(s['total_revenue'] or 0) for s in services]
        
        return JsonResponse({
            'labels': labels,
            'datasets': [{
                'label': 'Revenue (€)',
                'data': data,
                'backgroundColor': 'rgba(54, 162, 235, 0.8)',
                'borderColor': 'rgba(54, 162, 235, 1)',
                'borderWidth': 1
            }]
        })
    
    def _get_staff_data(self, start_date, end_date):
        """Staff performance"""
        staff_data = Visit.objects.filter(
            appointment__start__gte=start_date,
            appointment__start__lt=end_date
        ).values(
            'staff__name',
            'staff__username'
        ).annotate(
            visit_count=Count('id'),
            total_revenue=Sum('paid_amount')
        ).order_by('-total_revenue')[:10]
        
        labels = [s['staff__name'] or s['staff__username'] for s in staff_data]
        visits = [s['visit_count'] for s in staff_data]
        revenue = [float(s['total_revenue'] or 0) for s in staff_data]
        
        return JsonResponse({
            'labels': labels,
            'datasets': [
                {
                    'label': 'Visits',
                    'data': visits,
                    'backgroundColor': 'rgba(54, 162, 235, 0.8)',
                    'borderColor': 'rgba(54, 162, 235, 1)',
                    'borderWidth': 1,
                    'yAxisID': 'y'
                },
                {
                    'label': 'Revenue (€)',
                    'data': revenue,
                    'backgroundColor': 'rgba(75, 192, 192, 0.8)',
                    'borderColor': 'rgba(75, 192, 192, 1)',
                    'borderWidth': 1,
                    'yAxisID': 'y1'
                }
            ]
        })
    
    def _get_rooms_data(self, start_date, end_date):
        """Room utilization"""
        rooms = Appointment.objects.filter(
            start__gte=start_date,
            start__lt=end_date,
            room__isnull=False
        ).values(
            'room__name'
        ).annotate(
            count=Count('id')
        ).order_by('-count')
        
        labels = [r['room__name'] for r in rooms]
        data = [r['count'] for r in rooms]
        
        return JsonResponse({
            'labels': labels,
            'datasets': [{
                'data': data,
                'backgroundColor': [
                    'rgba(255, 99, 132, 0.8)',
                    'rgba(54, 162, 235, 0.8)',
                    'rgba(255, 206, 86, 0.8)',
                    'rgba(75, 192, 192, 0.8)',
                    'rgba(153, 102, 255, 0.8)',
                ],
                'borderWidth': 2
            }]
        })
    
    def _get_machines_data(self, start_date, end_date):
        """Machine utilization"""
        machines = Appointment.objects.filter(
            start__gte=start_date,
            start__lt=end_date,
            machine__isnull=False
        ).values(
            'machine__name'
        ).annotate(
            count=Count('id')
        ).order_by('-count')
        
        labels = [m['machine__name'] for m in machines]
        data = [m['count'] for m in machines]
        
        return JsonResponse({
            'labels': labels,
            'datasets': [{
                'label': 'Usage Count',
                'data': data,
                'backgroundColor': 'rgba(153, 102, 255, 0.8)',
                'borderColor': 'rgba(153, 102, 255, 1)',
                'borderWidth': 1
            }]
        })


# ═══════════════════════════════════════════════════════════════
# EXCEL EXPORT VIEW - YOUR ORIGINAL ONE
# ═══════════════════════════════════════════════════════════════

class ExportMonthlyReportView(LoginRequiredMixin, TemplateView):
    """Export monthly report to Excel - SUPPORTS DAILY, MONTHLY, AND DATE RANGE EXPORT"""
    
    def get(self, request, *args, **kwargs):
        if not EXCEL_AVAILABLE:
            return HttpResponse("Excel export not available. Install openpyxl.", status=500)
        
        # Check export type based on parameters
        selected_date_str = request.GET.get('date')
        selected_month = request.GET.get('month')
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if start_date_str and end_date_str:
            # DATE RANGE EXPORT
            try:
                start_date_only = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date_only = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                return HttpResponse("Invalid date format", status=400)
            
            start_date = timezone.make_aware(datetime.combine(start_date_only, datetime.min.time()))
            end_date = timezone.make_aware(datetime.combine(end_date_only, datetime.max.time()))
            filename_date = f"{start_date_only.strftime('%Y_%m_%d')}_to_{end_date_only.strftime('%Y_%m_%d')}"
            report_title = f"{start_date_only.strftime('%d %b %Y')} to {end_date_only.strftime('%d %b %Y')}"
            
        elif selected_date_str:
            # DAILY EXPORT
            try:
                selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
            except ValueError:
                selected_date = timezone.now().date()
            
            start_date = timezone.make_aware(datetime.combine(selected_date, datetime.min.time()))
            end_date = timezone.make_aware(datetime.combine(selected_date, datetime.max.time()))
            filename_date = selected_date.strftime('%Y_%m_%d')
            report_title = selected_date.strftime('%d %B %Y')
            
        else:
            # MONTHLY EXPORT (original logic)
            if selected_month:
                try:
                    selected_date = datetime.strptime(selected_month, '%Y-%m')
                except ValueError:
                    selected_date = timezone.now()
            else:
                selected_date = timezone.now()
            
            start_date = selected_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if start_date.month == 12:
                end_date = start_date.replace(year=start_date.year + 1, month=1)
            else:
                end_date = start_date.replace(month=start_date.month + 1)
            
            filename_date = selected_date.strftime('%Y_%m')
            report_title = selected_date.strftime('%B %Y')
        
        # Create workbook
        wb = Workbook()
        
        # Create existing sheets (pass report_title for proper labeling)
        self._create_summary_sheet(wb, start_date, end_date, report_title)
        self._create_visits_sheet(wb, start_date, end_date)
        self._create_services_sheet(wb, start_date, end_date)
        self._create_staff_sheet(wb, start_date, end_date)
        self._create_rooms_sheet(wb, start_date, end_date)
        self._create_machines_sheet(wb, start_date, end_date)
        
        # ✅ NEW: Create detailed appointment sheets
        self._create_cancelled_appointments_sheet(wb, start_date, end_date)
        self._create_no_show_appointments_sheet(wb, start_date, end_date)
        self._create_completed_appointments_sheet(wb, start_date, end_date)
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Report_{filename_date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    def _create_summary_sheet(self, wb, start_date, end_date, report_title):
        """Create summary statistics sheet"""
        ws = wb.active
        ws.title = "Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="5156BE", end_color="5156BE", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Title
        ws['A1'] = f"Report - {report_title}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        # Statistics
        visits = Visit.objects.filter(
            appointment__start__gte=start_date,
            appointment__start__lt=end_date
        )
        
        row = 3
        stats = [
            ('Total Visits', visits.count()),
            ('Total Revenue', f"€{visits.aggregate(Sum('charge_amount'))['charge_amount__sum'] or 0:.2f}"),
            ('Total Paid', f"€{visits.aggregate(Sum('paid_amount'))['paid_amount__sum'] or 0:.2f}"),
            ('Fully Paid Visits', visits.filter(paid_amount__gte=F('charge_amount')).count()),
            ('Partially Paid', visits.filter(paid_amount__gt=0, paid_amount__lt=F('charge_amount')).count()),
            ('Unpaid Visits', visits.filter(paid_amount=0).count()),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_visits_sheet(self, wb, start_date, end_date):
        """Create detailed visits sheet"""
        ws = wb.create_sheet("Visits Detail")
        
        # Headers
        headers = ['Date', 'Client', 'Service', 'Area', 'Staff', 'Charge', 'Paid', 'Balance', 'Payment Method']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5156BE", end_color="5156BE", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        visits = Visit.objects.filter(
            appointment__start__gte=start_date,
            appointment__start__lt=end_date
        ).select_related(
            'appointment__client',
            'appointment__service',
            'staff'
        ).order_by('appointment__start')
        
        for row, visit in enumerate(visits, 2):
            # Get staff name using flexible method
            staff = visit.staff
            if hasattr(staff, 'name') and staff.name:
                staff_name = staff.name
            elif hasattr(staff, 'get_full_name'):
                full_name = staff.get_full_name()
                staff_name = full_name if full_name.strip() else staff.username
            else:
                staff_name = staff.username
            
            ws.cell(row=row, column=1, value=visit.appointment.start.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=visit.appointment.client.full_name)
            ws.cell(row=row, column=3, value=visit.appointment.service.name)
            ws.cell(row=row, column=4, value=visit.area or '')
            ws.cell(row=row, column=5, value=staff_name)
            ws.cell(row=row, column=6, value=float(visit.charge_amount))
            ws.cell(row=row, column=7, value=float(visit.paid_amount))
            ws.cell(row=row, column=8, value=float(visit.charge_amount - visit.paid_amount))
            ws.cell(row=row, column=9, value=visit.get_payment_method_display() if visit.payment_method else '')
        
        # Auto-size columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_services_sheet(self, wb, start_date, end_date):
        """Create services breakdown sheet"""
        ws = wb.create_sheet("Services Breakdown")
        
        # Headers
        headers = ['Service', 'Count', 'Total Revenue', 'Avg per Visit']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5156BE", end_color="5156BE", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        services = Visit.objects.filter(
            appointment__start__gte=start_date,
            appointment__start__lt=end_date
        ).values(
            'appointment__service__name'
        ).annotate(
            count=Count('id'),
            revenue=Sum('charge_amount'),
            avg=Avg('charge_amount')
        ).order_by('-revenue')
        
        for row, service in enumerate(services, 2):
            ws.cell(row=row, column=1, value=service['appointment__service__name'])
            ws.cell(row=row, column=2, value=service['count'])
            ws.cell(row=row, column=3, value=float(service['revenue']))
            ws.cell(row=row, column=4, value=float(service['avg']))
        
        # Auto-size columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_staff_sheet(self, wb, start_date, end_date):
        """Create staff performance sheet"""
        ws = wb.create_sheet("Staff Performance")
        
        # Headers
        headers = ['Staff Member', 'Total Visits', 'Total Revenue', 'Avg per Visit']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5156BE", end_color="5156BE", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        staff_data = Visit.objects.filter(
            appointment__start__gte=start_date,
            appointment__start__lt=end_date
        ).values('staff__id').annotate(
            visit_count=Count('id'),
            revenue=Sum('charge_amount'),
            avg=Avg('charge_amount')
        ).order_by('-visit_count')
        
        row = 2
        for staff_item in staff_data:
            try:
                staff = User.objects.get(id=staff_item['staff__id'])
                
                # Try different methods to get staff name
                if hasattr(staff, 'name') and staff.name:
                    name = staff.name
                elif hasattr(staff, 'get_full_name'):
                    full_name = staff.get_full_name()
                    name = full_name if full_name.strip() else str(staff)
                elif hasattr(staff, 'first_name') and hasattr(staff, 'last_name'):
                    if staff.first_name and staff.last_name:
                        name = f"{staff.first_name} {staff.last_name}"
                    elif staff.first_name:
                        name = staff.first_name
                    else:
                        name = staff.username
                else:
                    name = str(staff)
                
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=staff_item['visit_count'])
                ws.cell(row=row, column=3, value=float(staff_item['revenue']))
                ws.cell(row=row, column=4, value=float(staff_item['avg']))
                row += 1
            except User.DoesNotExist:
                continue
        
        # Auto-size columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_rooms_sheet(self, wb, start_date, end_date):
        """Create room utilization sheet"""
        ws = wb.create_sheet("Room Utilization")
        
        # Headers
        headers = ['Room Name', 'Total Appointments', 'Completed', 'Cancelled', 'No Show']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5156BE", end_color="5156BE", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        rooms_data = Appointment.objects.filter(
            start__gte=start_date,
            start__lt=end_date
        ).values('room__name').annotate(
            total=Count('id'),
            completed=Count('id', filter=Q(status='completed')),
            cancelled=Count('id', filter=Q(status='cancelled')),
            no_show=Count('id', filter=Q(status='no_show'))
        ).order_by('-total')
        
        for row, room in enumerate(rooms_data, 2):
            ws.cell(row=row, column=1, value=room['room__name'])
            ws.cell(row=row, column=2, value=room['total'])
            ws.cell(row=row, column=3, value=room['completed'])
            ws.cell(row=row, column=4, value=room['cancelled'])
            ws.cell(row=row, column=5, value=room['no_show'])
        
        # Auto-size columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_machines_sheet(self, wb, start_date, end_date):
        """Create machine utilization sheet"""
        ws = wb.create_sheet("Machine Utilization")
        
        # Headers
        headers = ['Machine Name', 'Total Visits', 'Total Revenue', 'Avg per Visit']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5156BE", end_color="5156BE", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        machines_data = Visit.objects.filter(
            appointment__start__gte=start_date,
            appointment__start__lt=end_date,
            machine__isnull=False
        ).values('machine__name').annotate(
            visit_count=Count('id'),
            revenue=Sum('charge_amount'),
            avg=Avg('charge_amount')
        ).order_by('-visit_count')
        
        for row, machine in enumerate(machines_data, 2):
            ws.cell(row=row, column=1, value=machine['machine__name'])
            ws.cell(row=row, column=2, value=machine['visit_count'])
            ws.cell(row=row, column=3, value=float(machine['revenue']))
            ws.cell(row=row, column=4, value=float(machine['avg']))
        
        # Auto-size columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    # ========================================
    # ✅ NEW: 3 DETAILED APPOINTMENT SHEETS
    # ========================================
    
    def _create_cancelled_appointments_sheet(self, wb, start_date, end_date):
        """Create detailed list of cancelled appointments with client names"""
        ws = wb.create_sheet("Cancelled Appointments")
        
        # Headers
        headers = [
            'Date', 'Time', 'Client Name', 'Service', 'Staff', 
            'Room', 'Duration', 'Price', 'Cancellation Reason'
        ]
        
        # Style headers (RED for cancelled)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Get cancelled appointments with related data
        cancelled_appointments = Appointment.objects.filter(
            start__gte=start_date,
            start__lt=end_date,
            status='cancelled'
        ).select_related(
            'client', 'service', 'staff', 'room'
        ).order_by('start')
        
        # Write data
        row = 2
        for apt in cancelled_appointments:
            # Get staff name
            if hasattr(apt.staff, 'name') and apt.staff.name:
                staff_name = apt.staff.name
            else:
                staff_name = apt.staff.username
            
            # Calculate duration
            duration_minutes = int((apt.end - apt.start).total_seconds() / 60)
            
            # Get price
            price = float(apt.price_override) if apt.price_override else float(apt.service.default_price)
            
            ws.cell(row=row, column=1, value=apt.start.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=apt.start.strftime('%H:%M'))
            ws.cell(row=row, column=3, value=apt.client.full_name)
            ws.cell(row=row, column=4, value=apt.service.name)
            ws.cell(row=row, column=5, value=staff_name)
            ws.cell(row=row, column=6, value=apt.room.name)
            ws.cell(row=row, column=7, value=f"{duration_minutes} min")
            ws.cell(row=row, column=8, value=price)
            ws.cell(row=row, column=9, value=apt.notes or "")
            
            # Format price cell
            ws.cell(row=row, column=8).number_format = '€#,##0.00'
            
            row += 1
        
        # Add total count
        total_row = row + 1
        ws.cell(row=total_row, column=1, value="TOTAL:")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=f"{cancelled_appointments.count()} cancelled appointments")
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        
        # Calculate total lost revenue
        total_lost = sum(
            float(apt.price_override) if apt.price_override else float(apt.service.default_price)
            for apt in cancelled_appointments
        )
        ws.cell(row=total_row, column=7, value="Lost Revenue:")
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=total_lost)
        ws.cell(row=total_row, column=8).font = Font(bold=True, color="DC3545")
        ws.cell(row=total_row, column=8).number_format = '€#,##0.00'
        
        # Auto-size columns
        column_widths = {
            'A': 12, 'B': 10, 'C': 25, 'D': 30, 'E': 20,
            'F': 15, 'G': 12, 'H': 12, 'I': 40,
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_no_show_appointments_sheet(self, wb, start_date, end_date):
        """Create detailed list of no-show appointments with client names and phone"""
        ws = wb.create_sheet("No-Show Appointments")
        
        # Headers
        headers = [
            'Date', 'Time', 'Client Name', 'Client Phone', 'Service', 
            'Staff', 'Room', 'Price', 'Notes'
        ]
        
        # Style headers (ORANGE for no-show)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Get no-show appointments with related data
        no_show_appointments = Appointment.objects.filter(
            start__gte=start_date,
            start__lt=end_date,
            status='no_show'
        ).select_related(
            'client', 'service', 'staff', 'room'
        ).order_by('start')
        
        # Write data
        row = 2
        for apt in no_show_appointments:
            # Get staff name
            if hasattr(apt.staff, 'name') and apt.staff.name:
                staff_name = apt.staff.name
            else:
                staff_name = apt.staff.username
            
            # Get price
            price = float(apt.price_override) if apt.price_override else float(apt.service.default_price)
            
            ws.cell(row=row, column=1, value=apt.start.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=apt.start.strftime('%H:%M'))
            ws.cell(row=row, column=3, value=apt.client.full_name)
            ws.cell(row=row, column=4, value=apt.client.phone)
            ws.cell(row=row, column=5, value=apt.service.name)
            ws.cell(row=row, column=6, value=staff_name)
            ws.cell(row=row, column=7, value=apt.room.name)
            ws.cell(row=row, column=8, value=price)
            ws.cell(row=row, column=9, value=apt.notes or "")
            
            # Format price cell
            ws.cell(row=row, column=8).number_format = '€#,##0.00'
            
            # Highlight client phone for easy follow-up
            ws.cell(row=row, column=4).fill = PatternFill(
                start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
            )
            
            row += 1
        
        # Add total count
        total_row = row + 1
        ws.cell(row=total_row, column=1, value="TOTAL:")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=f"{no_show_appointments.count()} no-show appointments")
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        
        # Calculate total lost revenue
        total_lost = sum(
            float(apt.price_override) if apt.price_override else float(apt.service.default_price)
            for apt in no_show_appointments
        )
        ws.cell(row=total_row, column=7, value="Lost Revenue:")
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=total_lost)
        ws.cell(row=total_row, column=8).font = Font(bold=True, color="FFC107")
        ws.cell(row=total_row, column=8).number_format = '€#,##0.00'
        
        # Auto-size columns
        column_widths = {
            'A': 12, 'B': 10, 'C': 25, 'D': 15, 'E': 30,
            'F': 20, 'G': 15, 'H': 12, 'I': 40,
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_completed_appointments_sheet(self, wb, start_date, end_date):
        """Create detailed list of completed appointments with payment details"""
        ws = wb.create_sheet("Completed Appointments")
        
        # Headers
        headers = [
            'Date', 'Time', 'Client Name', 'Service', 'Staff', 
            'Room', 'Machine', 'Charged', 'Paid', 'Balance', 'Payment Status'
        ]
        
        # Style headers (GREEN for completed)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Get completed appointments with visits
        completed_appointments = Appointment.objects.filter(
            start__gte=start_date,
            start__lt=end_date,
            status='completed'
        ).select_related(
            'client', 'service', 'staff', 'room'
        ).prefetch_related('visit').order_by('start')
        
        # Write data
        row = 2
        total_charged = 0
        total_paid = 0
        
        for apt in completed_appointments:
            # Get staff name
            if hasattr(apt.staff, 'name') and apt.staff.name:
                staff_name = apt.staff.name
            else:
                staff_name = apt.staff.username
            
            # Get visit data if exists
            if hasattr(apt, 'visit'):
                visit = apt.visit
                charged = float(visit.charge_amount)
                paid = float(visit.paid_amount)
                balance = charged - paid
                
                # Determine payment status
                if paid >= charged:
                    payment_status = "Fully Paid"
                elif paid > 0:
                    payment_status = "Partially Paid"
                else:
                    payment_status = "Unpaid"
                
                machine_name = visit.machine.name if visit.machine else ""
            else:
                # No visit record
                price = float(apt.price_override) if apt.price_override else float(apt.service.default_price)
                charged = price
                paid = 0
                balance = charged
                payment_status = "No Visit Record"
                machine_name = ""
            
            ws.cell(row=row, column=1, value=apt.start.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=apt.start.strftime('%H:%M'))
            ws.cell(row=row, column=3, value=apt.client.full_name)
            ws.cell(row=row, column=4, value=apt.service.name)
            ws.cell(row=row, column=5, value=staff_name)
            ws.cell(row=row, column=6, value=apt.room.name)
            ws.cell(row=row, column=7, value=machine_name)
            ws.cell(row=row, column=8, value=charged)
            ws.cell(row=row, column=9, value=paid)
            ws.cell(row=row, column=10, value=balance)
            ws.cell(row=row, column=11, value=payment_status)
            
            # Format currency cells
            ws.cell(row=row, column=8).number_format = '€#,##0.00'
            ws.cell(row=row, column=9).number_format = '€#,##0.00'
            ws.cell(row=row, column=10).number_format = '€#,##0.00'
            
            # Color-code payment status
            status_cell = ws.cell(row=row, column=11)
            if payment_status == "Fully Paid":
                status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                status_cell.font = Font(color="155724")
            elif payment_status == "Partially Paid":
                status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                status_cell.font = Font(color="856404")
            elif payment_status == "Unpaid":
                status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                status_cell.font = Font(color="721C24")
            
            total_charged += charged
            total_paid += paid
            
            row += 1
        
        # Add totals row
        total_row = row + 1
        ws.cell(row=total_row, column=1, value="TOTALS:")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=f"{completed_appointments.count()} completed appointments")
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        
        ws.cell(row=total_row, column=7, value="Total Charged:")
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=total_charged)
        ws.cell(row=total_row, column=8).font = Font(bold=True)
        ws.cell(row=total_row, column=8).number_format = '€#,##0.00'
        
        ws.cell(row=total_row + 1, column=7, value="Total Paid:")
        ws.cell(row=total_row + 1, column=7).font = Font(bold=True)
        ws.cell(row=total_row + 1, column=8, value=total_paid)
        ws.cell(row=total_row + 1, column=8).font = Font(bold=True, color="28A745")
        ws.cell(row=total_row + 1, column=8).number_format = '€#,##0.00'
        
        ws.cell(row=total_row + 2, column=7, value="Outstanding:")
        ws.cell(row=total_row + 2, column=7).font = Font(bold=True)
        ws.cell(row=total_row + 2, column=8, value=total_charged - total_paid)
        ws.cell(row=total_row + 2, column=8).font = Font(bold=True, color="DC3545")
        ws.cell(row=total_row + 2, column=8).number_format = '€#,##0.00'
        
        # Auto-size columns
        column_widths = {
            'A': 12, 'B': 10, 'C': 25, 'D': 30, 'E': 20,
            'F': 15, 'G': 20, 'H': 12, 'I': 12, 'J': 12, 'K': 15,
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'


class AnalyticsDebugView(LoginRequiredMixin, TemplateView):
    """Debug view to help troubleshoot staff, room, and machine data"""
    template_name = 'analytics/debug.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        # Get recent visits
        context['recent_visits'] = Visit.objects.select_related(
            'staff', 'appointment', 'machine'
        ).order_by('-created_at')[:10]
        
        # Get User model field information
        user_fields = []
        for field in User._meta.get_fields():
            try:
                user_fields.append({
                    'name': field.name,
                    'type': field.__class__.__name__
                })
            except:
                pass
        context['user_fields'] = user_fields
        
        # Test staff performance query
        from datetime import datetime, timedelta
        end_date = timezone.now()
        start_date = end_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        staff_data = Visit.objects.filter(
            appointment__start__gte=start_date,
            appointment__start__lt=end_date
        ).values('staff__id').annotate(
            visit_count=Count('id'),
            revenue=Sum('charge_amount')
        ).order_by('-visit_count')[:10]
        
        staff_stats = []
        for s in staff_data:
            try:
                staff = User.objects.get(id=s['staff__id'])
                
                # Try different methods to get staff name
                if hasattr(staff, 'name') and staff.name:
                    name = staff.name
                elif hasattr(staff, 'get_full_name'):
                    full_name = staff.get_full_name()
                    name = full_name if full_name.strip() else str(staff)
                else:
                    name = str(staff)
                
                staff_stats.append({
                    'name': name,
                    'visit_count': s['visit_count'],
                    'revenue': s['revenue']
                })
            except User.DoesNotExist:
                continue
        
        context['staff_stats'] = staff_stats
        
        # Get recent appointments
        context['recent_appointments'] = Appointment.objects.select_related(
            'room'
        ).order_by('-created_at')[:10]
        
        # Get visits with machines
        context['visits_with_machines'] = Visit.objects.filter(
            machine__isnull=False
        ).select_related('machine').order_by('-created_at')[:10]
        
        # Summary statistics
        context['total_visits'] = Visit.objects.count()
        context['visits_with_staff'] = Visit.objects.filter(staff__isnull=False).count()
        context['visits_with_machines_count'] = Visit.objects.filter(machine__isnull=False).count()
        context['total_appointments'] = Appointment.objects.count()
        context['unique_staff'] = Visit.objects.values('staff').distinct().count()
        context['unique_rooms'] = Appointment.objects.values('room').distinct().count()
        context['unique_machines'] = Visit.objects.filter(machine__isnull=False).values('machine').distinct().count()
        
        return context