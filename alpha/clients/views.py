from django.shortcuts import render

# Create your views here.
"""
Clients Views
Place this at: alpha/clients/views.py
"""
from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import Q, Count, Sum
from django.contrib import messages
from django.utils.translation import gettext as _
from .models import Client


class ClientListView(LoginRequiredMixin, ListView):
    model = Client
    template_name = 'clients/client_list.html'
    context_object_name = 'clients'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = Client.objects.annotate(
            appointment_count=Count('appointments'),
            visit_count=Count('appointments__visit')
        ).all()
        
        skin_type = self.request.GET.get('skin_type')
        if skin_type:
            queryset = queryset.filter(skin_type=skin_type)
        
        hair_color = self.request.GET.get('hair_color')
        if hair_color:
            queryset = queryset.filter(hair_color=hair_color)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search) |
                Q(phone__icontains=search) |
                Q(email__icontains=search) |
                Q(notes__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_count'] = Client.objects.count()
        context['with_email_count'] = Client.objects.exclude(email='').count()
        context['current_skin_type'] = self.request.GET.get('skin_type', '')
        context['current_hair_color'] = self.request.GET.get('hair_color', '')
        context['current_search'] = self.request.GET.get('search', '')
        context['skin_type_choices'] = Client.SKIN_TYPE_CHOICES
        context['hair_color_choices'] = Client.HAIR_COLOR_CHOICES
        return context


class ClientCreateView(LoginRequiredMixin, CreateView):
    model = Client
    template_name = 'clients/client_form.html'
    fields = [
        'full_name', 'phone', 'email', 'birth_date', 
        'skin_type', 'hair_color', 'notes',
        'receive_booking_sms', 'receive_booking_email',
        'receive_reminder_sms', 'receive_reminder_email'
    ]
    success_url = reverse_lazy('clients:list')
    
    def form_valid(self, form):
        messages.success(self.request, _('Client created successfully!'))
        return super().form_valid(form)


class ClientDetailView(LoginRequiredMixin, DetailView):
    model = Client
    template_name = 'clients/client_detail.html'
    context_object_name = 'client'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        appointments = self.object.appointments.select_related(
            'service', 'staff', 'room', 'machine'
        ).order_by('-start')[:10]
        
        context['recent_appointments'] = appointments
        context['total_appointments'] = self.object.appointments.count()
        context['completed_visits'] = self.object.appointments.filter(status='completed').count()
        
        from alpha.visits.models import Visit
        visits = Visit.objects.filter(
            appointment__client=self.object
        ).select_related(
            'appointment',
            'appointment__service',
            'staff',
            'machine'
        ).order_by('-appointment__start')[:10]
        
        context['recent_visits'] = visits
        context['total_visits'] = Visit.objects.filter(appointment__client=self.object).count()
        
        revenue_stats = Visit.objects.filter(
            appointment__client=self.object
        ).aggregate(
            total_charged=Sum('charge_amount'),
            total_paid=Sum('paid_amount')
        )
        context['total_revenue'] = revenue_stats['total_paid'] or 0
        context['total_charged'] = revenue_stats['total_charged'] or 0
        context['consents'] = self.object.consents.all().order_by('-accepted_at')
        
        return context


class ClientUpdateView(LoginRequiredMixin, UpdateView):
    model = Client
    template_name = 'clients/client_form.html'
    fields = [
        'full_name', 'phone', 'email', 'birth_date', 
        'skin_type', 'hair_color', 'notes',
        'receive_booking_sms', 'receive_booking_email',
        'receive_reminder_sms', 'receive_reminder_email'
    ]
    
    def get_success_url(self):
        messages.success(self.request, _('Client updated successfully!'))
        return reverse_lazy('clients:detail', kwargs={'pk': self.object.pk})


class ClientDeleteView(LoginRequiredMixin, DeleteView):
    model = Client
    template_name = 'clients/client_confirm_delete.html'
    success_url = reverse_lazy('clients:list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Client deleted successfully.'))
        return super().delete(request, *args, **kwargs)

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class ClientExportExcelView(LoginRequiredMixin, ListView):
    """Export clients to Excel with filters"""
    model = Client
    
    def get_queryset(self):
        """
        Get clients with the same filters as ClientListView
        """
        queryset = Client.objects.annotate(
            appointment_count=Count('appointments'),
            visit_count=Count('appointments__visit')
        ).all()
        
        # Apply same filters as list view
        skin_type = self.request.GET.get('skin_type')
        if skin_type:
            queryset = queryset.filter(skin_type=skin_type)
        
        hair_color = self.request.GET.get('hair_color')
        if hair_color:
            queryset = queryset.filter(hair_color=hair_color)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search) |
                Q(phone__icontains=search) |
                Q(email__icontains=search) |
                Q(notes__icontains=search)
            )
        
        return queryset.order_by('full_name')
    
    def get(self, request, *args, **kwargs):
        """Generate and return Excel file"""
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clients"
        
        # Define column headers
        headers = [
            'ID',
            'Full Name',
            'Phone',
            'Email',
            'Birth Date',
            'Age',
            'Skin Type',
            'Hair Color',
            'Total Appointments',
            'Total Visits',
            'Notes',
            'Created At',
        ]
        
        # Style for header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Get clients
        clients = self.get_queryset()
        
        # Write data rows
        for row_num, client in enumerate(clients, 2):
            # Calculate age if birth_date exists
            age = ""
            if client.birth_date:
                today = datetime.now().date()
                age = today.year - client.birth_date.year
                if today.month < client.birth_date.month or (
                    today.month == client.birth_date.month and today.day < client.birth_date.day
                ):
                    age -= 1
            
            # Get display values
            skin_type_display = client.get_skin_type_display() if client.skin_type else ""
            hair_color_display = client.get_hair_color_display() if client.hair_color else ""
            
            row_data = [
                client.id,
                client.full_name,
                client.phone,
                client.email or "",
                client.birth_date.strftime('%d/%m/%Y') if client.birth_date else "",
                age,
                skin_type_display,
                hair_color_display,
                client.appointment_count,
                client.visit_count,
                client.notes or "",
                client.created_at.strftime('%d/%m/%Y %H:%M') if hasattr(client, 'created_at') else "",
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = thin_border
                
                # Center align numbers
                if col_num in [1, 6, 9, 10]:  # ID, Age, Appointments, Visits
                    cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        column_widths = {
            'A': 8,   # ID
            'B': 25,  # Full Name
            'C': 15,  # Phone
            'D': 30,  # Email
            'E': 12,  # Birth Date
            'F': 8,   # Age
            'G': 15,  # Skin Type
            'H': 15,  # Hair Color
            'I': 12,  # Appointments
            'J': 10,  # Visits
            'K': 40,  # Notes
            'L': 18,  # Created At
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'clients_export_{timestamp}.xlsx'
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        
        return response